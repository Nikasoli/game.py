from openpyxl import Workbook
import os
os.system('cls')

wb=Workbook()
ws=wb.active
ws.title='amar'

for i in range(1,11):
    ws[f'a{i}']=i
file=open('test.txt')
color=file.read()
colors=color.split('\n')
for i in colors:
    ws[f'b{colors.index(i)+1}']=i



wb.save('test.xlsx')